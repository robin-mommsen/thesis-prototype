#!/usr/bin/env python3
import argparse
import statistics
import sys
from scipy.stats import wilcoxon
from pathlib import Path
import krippendorff
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DIMENSIONS = ["Aufgabenpassung", "Materialintegration", "Umsetzbarkeit", "Kreativität"]
SCORE_COLUMNS = {
    "Aufgabenpassung": 4,
    "Materialintegration": 5,
    "Umsetzbarkeit": 6,
    "Kreativität": 7,
}
MODES = [("Ohne RAG", "baseline"), ("Mit RAG", "rag")]
GROUPS = [("Gesamt", None), ("Vage", "vage"), ("Mittel", "mittel"), ("Konkret", "konkret")]

COLOR_HEADER = "1F3864"
COLOR_ANALYSIS = "1A5276"
COLOR_DIFF = "EAF2F8"
COLOR_WARNING = "FDEDEC"
COLOR_SIG = "D5F5E3"


def thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def header_font():
    return Font(color="FFFFFF", bold=True, size=10)


def normal_font():
    return Font(size=10)


def bold_font():
    return Font(bold=True, size=10)


def read_mapping(workbook_path):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if "Mapping" not in wb.sheetnames:
        raise ValueError(f"{workbook_path} has no Mapping sheet")

    ws = wb["Mapping"]
    mapping = {}
    for row in range(2, ws.max_row + 1):
        response_id = ws.cell(row=row, column=1).value
        if not response_id or response_id == "EXCLUDED":
            continue
        if not str(response_id)[0:1].isalpha() or not str(response_id)[1:].isdigit():
            continue
        mapping[str(response_id)] = {
            "response_id": str(response_id),
            "source_id": ws.cell(row=row, column=2).value,
            "input_type": ws.cell(row=row, column=3).value,
            "mode": ws.cell(row=row, column=4).value,
            "run": ws.cell(row=row, column=5).value,
            "duration_ms": ws.cell(row=row, column=6).value,
        }

    if not mapping:
        raise ValueError(f"{workbook_path} contains no scorable Mapping rows")
    return mapping


def read_scores(workbook_path, rater_name):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if "Scoring" not in wb.sheetnames:
        raise ValueError(f"{workbook_path} has no Scoring sheet")

    ws = wb["Scoring"]
    rows = []
    for row in range(3, ws.max_row + 1):
        response_id = ws.cell(row=row, column=1).value
        if not response_id:
            continue

        scores = {}
        for dimension, column in SCORE_COLUMNS.items():
            value = ws.cell(row=row, column=column).value
            if value is None or value == "":
                scores[dimension] = None
                continue
            try:
                parsed = int(value)
            except (TypeError, ValueError):
                parsed = None
            scores[dimension] = parsed

        rows.append({
            "response_id": str(response_id),
            "rater": rater_name,
            "scores": scores,
        })
    return rows


def validate_scores(all_scores, mapping, expected_raters, allow_incomplete):
    errors = []
    grouped = {response_id: [] for response_id in mapping}

    for score_row in all_scores:
        response_id = score_row["response_id"]
        if response_id not in mapping:
            errors.append(f"Unknown response ID {response_id} in rater {score_row['rater']}")
            continue
        grouped[response_id].append(score_row)

        for dimension, value in score_row["scores"].items():
            if value is None:
                errors.append(f"Missing score: {response_id}, {score_row['rater']}, {dimension}")
            elif value < 1 or value > 4:
                errors.append(f"Invalid score {value}: {response_id}, {score_row['rater']}, {dimension}")

    for response_id, rows in grouped.items():
        raters = {row["rater"] for row in rows}
        if len(raters) != expected_raters:
            errors.append(f"{response_id} has {len(raters)} rater(s), expected {expected_raters}")

    if errors and not allow_incomplete:
        preview = "\n".join(f"  - {error}" for error in errors[:20])
        more = "" if len(errors) <= 20 else f"\n  ... and {len(errors) - 20} more"
        raise ValueError(f"Ratings are incomplete or invalid:\n{preview}{more}")

    return grouped, errors


def aggregate_response_scores(grouped, mapping):
    aggregated = []

    for response_id in sorted(mapping):
        score_rows = grouped.get(response_id, [])
        item = dict(mapping[response_id])
        item["rater_count"] = len({row["rater"] for row in score_rows})

        dimension_means = []
        for dimension in DIMENSIONS:
            values = [
                row["scores"][dimension]
                for row in score_rows
                if row["scores"].get(dimension) is not None
            ]
            mean = statistics.mean(values) if values else None
            stdev = statistics.stdev(values) if len(values) > 1 else None
            item[f"{dimension}_mean"] = mean
            item[f"{dimension}_stdev"] = stdev
            if mean is not None:
                dimension_means.append(mean)

        item["overall_mean"] = statistics.mean(dimension_means) if dimension_means else None
        aggregated.append(item)

    return aggregated


def krippendorff_alpha_for_dimension(grouped, dimension):
    units = list(grouped.values())

    all_raters = []
    for score_rows in units:
        for row in score_rows:
            if row["rater"] not in all_raters:
                all_raters.append(row["rater"])

    data = []
    for rater in all_raters:
        rater_row = []
        for score_rows in units:
            score = next(
                (r["scores"].get(dimension) for r in score_rows if r["rater"] == rater),
                np.nan
            )
            rater_row.append(score if score is not None else np.nan)
        data.append(rater_row)

    try:
        alpha = krippendorff.alpha(
            reliability_data=data,
            level_of_measurement="ordinal",
            value_domain=[1, 2, 3, 4],
        )
    except Exception as e:
        return {
            "dimension": dimension,
            "alpha": None,
            "observed_disagreement": None,
            "expected_disagreement": None,
            "units": len(units),
            "ratings": sum(1 for row in data for v in row if not np.isnan(v)),
            "note": f"Nicht berechenbar: {e}",
        }

    return {
        "dimension": dimension,
        "alpha": alpha,
        "observed_disagreement": None,
        "expected_disagreement": None,
        "units": len(units),
        "ratings": sum(1 for row in data for v in row if not np.isnan(v)),
        "note": "Krippendorffs Alpha, ordinale Distanzfunktion, value_domain=[1,2,3,4]",
    }


def calculate_reliability(grouped):
    return [krippendorff_alpha_for_dimension(grouped, dimension) for dimension in DIMENSIONS]


def calculate_wilcoxon(grouped, mapping):
    def build_prompt_level_pairs(dimension, source_ids=None):
        pairs_by_source = {}

        for response_id, score_rows in grouped.items():
            meta = mapping[response_id]
            source_id = meta["source_id"]
            mode = meta["mode"]

            if source_ids is not None and source_id not in source_ids:
                continue

            scores = []
            if dimension == "overall":
                for row in score_rows:
                    vals = [
                        row["scores"].get(dim)
                        for dim in DIMENSIONS
                        if row["scores"].get(dim) is not None
                    ]
                    if vals:
                        scores.append(statistics.mean(vals))
            else:
                for row in score_rows:
                    score = row["scores"].get(dimension)
                    if score is not None:
                        scores.append(score)

            if scores:
                pairs_by_source.setdefault(source_id, {})[mode] = statistics.mean(scores)

        baseline_vals, rag_vals = [], []
        for source_id, modes in pairs_by_source.items():
            if "baseline" in modes and "rag" in modes:
                baseline_vals.append(modes["baseline"])
                rag_vals.append(modes["rag"])

        return baseline_vals, rag_vals

    def run_tests(source_ids=None):
        results = {}

        for dimension in DIMENSIONS + ["overall"]:
            baseline_vals, rag_vals = build_prompt_level_pairs(dimension, source_ids)

            differences = [r - b for b, r in zip(baseline_vals, rag_vals)]
            n_pairs = len(baseline_vals)

            if all(d == 0 for d in differences):
                results[dimension] = {
                    "n_pairs": n_pairs,
                    "W": None,
                    "p": None,
                    "r": None,
                    "mean_diff": 0.0,
                    "note": "Alle Differenzen sind 0"
                }
                continue

            try:
                result = wilcoxon(
                    baseline_vals,
                    rag_vals,
                    alternative="two-sided",
                    method="auto"
                )

                stat = result.statistic
                p_value = result.pvalue

                if hasattr(result, "zstatistic"):
                    z = abs(result.zstatistic)
                    r_effect = z / (n_pairs ** 0.5)
                else:
                    r_effect = None

                results[dimension] = {
                    "n_pairs": n_pairs,
                    "W": stat,
                    "p": p_value,
                    "r": r_effect,
                    "mean_diff": statistics.mean(differences),
                    "note": "Wilcoxon auf Prompt-Ebene: Raterwerte je Prompt aggregiert"
                }

            except Exception as e:
                results[dimension] = {
                    "n_pairs": n_pairs,
                    "W": None,
                    "p": None,
                    "r": None,
                    "mean_diff": None,
                    "note": f"Fehler: {e}"
                }

        return results

    output = {"Gesamt": run_tests(source_ids=None)}

    for group_label, input_type in [
        ("Vage", "vage"),
        ("Mittel", "mittel"),
        ("Konkret", "konkret")
    ]:
        source_ids = {
            meta["source_id"]
            for meta in mapping.values()
            if meta["input_type"] == input_type
        }
        output[group_label] = run_tests(source_ids=source_ids)

    return output


def write_sheet_header(ws, headers, row=1, fill=COLOR_HEADER):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font()
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()


def write_aggregated_sheet(wb, aggregated):
    ws = wb.active
    ws.title = "Aggregierte Antworten"

    headers = [
        "Antwort-ID", "Original-ID", "Typ", "Modus", "Run", "Rater n",
        "Aufgabenpassung MW", "Aufgabenpassung SD",
        "Materialintegration MW", "Materialintegration SD",
        "Umsetzbarkeit MW", "Umsetzbarkeit SD",
        "Kreativität MW", "Kreativität SD",
        "Gesamt MW",
    ]
    write_sheet_header(ws, headers)

    for row_idx, item in enumerate(aggregated, start=2):
        values = [
            item["response_id"],
            item["source_id"],
            item["input_type"],
            item["mode"],
            item["run"],
            item["rater_count"],
            item["Aufgabenpassung_mean"],
            item["Aufgabenpassung_stdev"],
            item["Materialintegration_mean"],
            item["Materialintegration_stdev"],
            item["Umsetzbarkeit_mean"],
            item["Umsetzbarkeit_stdev"],
            item["Kreativität_mean"],
            item["Kreativität_stdev"],
            item["overall_mean"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border()
            cell.font = normal_font()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if isinstance(value, float):
                cell.number_format = "0.00"

    widths = [12, 16, 12, 12, 8, 10, 20, 18, 22, 20, 18, 18, 18, 18, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:O{len(aggregated) + 1}"


def condition_mean(aggregated, mode, input_type, dimension):
    values = []
    for item in aggregated:
        if item["mode"] != mode:
            continue
        if input_type and item["input_type"] != input_type:
            continue
        value = item[f"{dimension}_mean"]
        if value is not None:
            values.append(value)
    return statistics.mean(values) if values else None


def condition_count(aggregated, mode, input_type):
    return sum(
        1
        for item in aggregated
        if item["mode"] == mode and (not input_type or item["input_type"] == input_type)
    )


def write_analysis_sheet(wb, aggregated):
    ws = wb.create_sheet("Auswertung")
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

    title = ws.cell(row=1, column=1, value="Aggregierte Auswertung: RAG vs. Baseline")
    title.font = Font(color="FFFFFF", bold=True, size=13)
    title.fill = PatternFill("solid", fgColor=COLOR_ANALYSIS)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:H1")

    note = ws.cell(
        row=2,
        column=1,
        value=(
            "Die Werte basieren zuerst auf Mittelwerten pro Antwort-ID und Kriterium "
            "über alle Rater. Erst danach werden Mittelwerte pro Bedingung berechnet"
        ),
    )
    note.font = Font(size=9, italic=True, color="666666")
    ws.merge_cells("A2:H2")

    headers = ["Gruppe", "Bedingung", "n", *DIMENSIONS, "Mittelwert"]
    write_sheet_header(ws, headers, row=4, fill="2E4057")

    current_row = 5
    last_rows = {}
    for group_label, input_type in GROUPS:
        for mode_label, mode in MODES:
            values = [condition_mean(aggregated, mode, input_type, dimension) for dimension in DIMENSIONS]
            mean_values = [value for value in values if value is not None]
            overall = statistics.mean(mean_values) if mean_values else None
            row_values = [group_label, mode_label, condition_count(aggregated, mode, input_type), *values, overall]
            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = thin_border()
                cell.font = normal_font()
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(value, float):
                    cell.number_format = "0.00"
            last_rows[(group_label, mode)] = current_row
            current_row += 1

        baseline_row = last_rows[(group_label, "baseline")]
        rag_row = last_rows[(group_label, "rag")]
        ws.cell(row=current_row, column=1, value=group_label)
        ws.cell(row=current_row, column=2, value="Delta RAG - Baseline")
        for col_idx in range(4, 9):
            rag_value = ws.cell(row=rag_row, column=col_idx).value
            baseline_value = ws.cell(row=baseline_row, column=col_idx).value
            delta = rag_value - baseline_value if isinstance(rag_value, (int, float)) and isinstance(baseline_value,
                                                                                                     (int,
                                                                                                      float)) else None
            cell = ws.cell(row=current_row, column=col_idx, value=delta)
            if isinstance(delta, float):
                cell.number_format = "0.00"
        for col_idx in range(1, 9):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border()
            cell.font = bold_font()
            cell.fill = PatternFill("solid", fgColor=COLOR_DIFF)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1


def write_wilcoxon_sheet(wb, wilcoxon_results):
    """Eigenes Sheet mit Wilcoxon-Testergebnissen. Grün = p < 0.05."""
    ws = wb.create_sheet("Wilcoxon-Test")

    title = ws.cell(row=1, column=1,
                    value="Wilcoxon-Vorzeichen-Rang-Test: RAG vs. Baseline")
    title.font = Font(color="FFFFFF", bold=True, size=12)
    title.fill = PatternFill("solid", fgColor="1A5276")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 22

    headers = [
        "Gruppe", "Kriterium", "n Paare",
        "Mittl. Differenz\n(RAG − Baseline)",
        "W", "p-Wert", "signifikant?"
    ]

    write_sheet_header(ws, headers, row=4, fill="2E4057")
    ws.row_dimensions[4].height = 30

    dim_labels = {d: d for d in DIMENSIONS}
    dim_labels["overall"] = "Gesamt (alle Kriterien)"

    current_row = 5
    for group in ["Gesamt", "Vage", "Mittel", "Konkret"]:
        if group not in wilcoxon_results:
            continue

        first_in_group = True
        for dimension in DIMENSIONS + ["overall"]:
            r = wilcoxon_results[group].get(dimension, {})

            p_val = r.get("p")
            is_sig = p_val is not None and p_val < 0.05
            sig_label = "ja" if is_sig else ("nein" if p_val is not None else "–")

            row_vals = [
                group if first_in_group else "",
                dim_labels[dimension],
                r.get("n_pairs"),
                r.get("mean_diff"),
                r.get("W"),
                p_val,
                sig_label
            ]
            first_in_group = False

            for col_idx, value in enumerate(row_vals, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = thin_border()
                cell.font = normal_font()
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                if col_idx == 4 and isinstance(value, float):
                    cell.number_format = "+0.000;-0.000;0.000"
                elif col_idx == 5 and isinstance(value, float):
                    cell.number_format = "0.000"
                elif col_idx == 6 and isinstance(value, float):
                    cell.number_format = "0.000"

                if is_sig:
                    cell.fill = PatternFill("solid", fgColor=COLOR_SIG)

            current_row += 1

        current_row += 1

    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)

    widths = [12, 26, 10, 20, 10, 10, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_raw_sheet(wb, all_scores):
    ws = wb.create_sheet("Raterdaten")
    headers = ["Antwort-ID", "Rater", *DIMENSIONS]
    write_sheet_header(ws, headers)

    for row_idx, row in enumerate(all_scores, start=2):
        values = [
            row["response_id"],
            row["rater"],
            row["scores"]["Aufgabenpassung"],
            row["scores"]["Materialintegration"],
            row["scores"]["Umsetzbarkeit"],
            row["scores"]["Kreativität"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border()
            cell.font = normal_font()
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [14, 28, 18, 22, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def write_reliability_sheet(wb, reliability):
    ws = wb.create_sheet("Inter-Rater-Reliabilität")
    headers = [
        "Kriterium",
        "Krippendorffs Alpha",
        "Antworten n",
        "Ratings n",
        "Hinweis",
    ]
    write_sheet_header(ws, headers)

    for row_idx, item in enumerate(reliability, start=2):
        values = [
            item["dimension"],
            item["alpha"],
            item["units"],
            item["ratings"],
            item["note"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border()
            cell.font = normal_font()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(value, float):
                cell.number_format = "0.000"

    note_row = len(reliability) + 4
    ws.cell(row=note_row, column=1).font = bold_font()
    ws.cell(row=note_row, column=2).alignment = Alignment(wrap_text=True)

    widths = [24, 22, 24, 24, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_validation_sheet(wb, errors):
    ws = wb.create_sheet("Validierung")
    headers = ["Typ", "Hinweis"]
    write_sheet_header(ws, headers)

    rows = [("OK", "Alle erwarteten Raterbewertungen sind vorhanden und gültig")] if not errors else [
        ("Warnung", error) for error in errors
    ]
    for row_idx, values in enumerate(rows, start=2):
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border()
            cell.font = normal_font()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if values[0] == "Warnung":
                cell.fill = PatternFill("solid", fgColor=COLOR_WARNING)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 90


def main():
    parser = argparse.ArgumentParser(description="Mehrere Rater-Excel-Dateien aggregieren")
    parser.add_argument("files", nargs="+", help="Ausgefüllte Scoring-Dateien der Rater")
    parser.add_argument("--master", required=True, help="Master-Excel-Datei mit Mapping-Sheet")
    parser.add_argument("--output", default="aggregated_rater_scores.xlsx", help="Output-Excel-Datei")
    parser.add_argument("--allow-incomplete", action="store_true",
                        help="Auch bei fehlenden Bewertungen eine Datei erzeugen")
    args = parser.parse_args()

    input_files = [Path(file) for file in args.files]
    missing_files = [file for file in input_files if not file.exists()]
    if missing_files:
        print("Error: missing file(s):")
        for file in missing_files:
            print(f"  - {file}")
        sys.exit(1)

    try:
        mapping = read_mapping(Path(args.master))
        all_scores = []
        for file in input_files:
            all_scores.extend(read_scores(file, file.stem))

        grouped, errors = validate_scores(
            all_scores,
            mapping,
            expected_raters=len(input_files),
            allow_incomplete=args.allow_incomplete,
        )
        aggregated = aggregate_response_scores(grouped, mapping)
        reliability = calculate_reliability(grouped)
        wilcoxon_results = calculate_wilcoxon(grouped, mapping)
    except Exception as exc:
        print(f"Error: {exc}")
        sys.exit(1)

    output_path = Path(args.output)
    wb = openpyxl.Workbook()
    write_aggregated_sheet(wb, aggregated)
    write_analysis_sheet(wb, aggregated)
    write_wilcoxon_sheet(wb, wilcoxon_results)
    write_raw_sheet(wb, all_scores)
    write_reliability_sheet(wb, reliability)
    write_validation_sheet(wb, errors)
    wb.save(output_path)

    print(f"\nCreated: {output_path}")
    print(f"  Responses: {len(aggregated)}")
    print(f"  Rater files: {len(input_files)}")
    print(f"  Raw rating rows: {len(all_scores)}")
    if errors:
        print(f"  Warnings: {len(errors)}")
    else:
        print("  Validation: OK")
    print("  Interrater reliability: Krippendorff's alpha per criterion")
    if wilcoxon_results:
        print("  Wilcoxon test: completed")


if __name__ == "__main__":
    main()
