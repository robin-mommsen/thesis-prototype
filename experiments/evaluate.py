#!/usr/bin/env python3
import argparse
import json
import random
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


DIMENSIONS = ["Aufgabenpassung", "Materialintegration", "Umsetzbarkeit", "Kreativität"]
MODES = [("Ohne RAG", "baseline"), ("Mit RAG", "rag")]

COLOR_HEADER = "1F3864"
COLOR_WARNING = "FDEDEC"


def thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def header_font():
    return Font(color="FFFFFF", bold=True, size=10)


def normal_font():
    return Font(size=10)


def bold_font():
    return Font(bold=True, size=10)


def get_runs(item, mode):
    if mode == "rag":
        if "rag_runs" in item:
            return item["rag_runs"] or []
        if "rag" in item:
            return [item["rag"]]
    else:
        if "baseline_runs" in item:
            return item["baseline_runs"] or []
        if "baseline" in item:
            return [item["baseline"]]
    return []


def collect_responses(results, seed, include_errors=False):
    rows = []

    for item_idx, item in enumerate(results, start=1):
        source_id = item.get("id", f"item_{item_idx:03d}")
        input_type = item.get("type", "")
        user_input = item.get("input", "")

        for mode in ["rag", "baseline"]:
            runs = get_runs(item, mode)
            for run_idx, run_data in enumerate(runs, start=1):
                success = bool(run_data.get("success"))

                if success:
                    response_text = run_data.get("response", "")
                else:
                    response_text = f"[Fehler: {run_data.get('error', '')}]"
                    if not include_errors:
                        rows.append({
                            "source_id": source_id,
                            "input_type": input_type,
                            "input": user_input,
                            "mode": mode,
                            "run": run_idx,
                            "response": response_text,
                            "duration_ms": run_data.get("duration_ms"),
                            "success": success,
                            "include_in_scoring": False,
                        })
                        continue

                rows.append({
                    "source_id": source_id,
                    "input_type": input_type,
                    "input": user_input,
                    "mode": mode,
                    "run": run_idx,
                    "response": response_text,
                    "duration_ms": run_data.get("duration_ms"),
                    "success": success,
                    "include_in_scoring": True,
                })

    scorable_rows = [row for row in rows if row["include_in_scoring"]]
    random.Random(seed).shuffle(scorable_rows)

    response_counter = 1
    for row in scorable_rows:
        row["response_id"] = f"A{response_counter:03d}"
        response_counter += 1

    for row in rows:
        if not row.get("response_id"):
            row["response_id"] = "EXCLUDED"

    return rows


def create_mapping_sheet(ws, rows, seed, include_errors):
    ws.title = "Mapping"

    headers = [
        "Antwort-ID",
        "Original-ID",
        "Typ",
        "Modus",
        "Run",
        "Dauer ms",
        "Erfolgreich",
        "Im Scoring enthalten",
        "Random Seed",
        "Fehler/Antwortauszug",
    ]

    widths = [14, 18, 14, 14, 8, 12, 12, 20, 14, 60]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font()
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border()

    for row_idx, item in enumerate(rows, start=2):
        excerpt = item["response"][:250] if item["response"] else ""
        values = [
            item["response_id"],
            item["source_id"],
            item["input_type"],
            item["mode"],
            item["run"],
            item["duration_ms"],
            item["success"],
            item["include_in_scoring"],
            seed if row_idx == 2 else "",
            excerpt,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border()
            cell.font = normal_font()
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if not item["include_in_scoring"]:
                cell.fill = PatternFill("solid", fgColor=COLOR_WARNING)

    note_row = len(rows) + 4
    ws.cell(row=note_row, column=1, value="Hinweis")
    ws.cell(
        row=note_row,
        column=2,
        value=(
            "Diese Datei ist die Master-Datei und verbleibt beim Versuchsleiter. "
            "Sie enthält die Zuordnung zwischen anonymisierter Antwort-ID und Bedingung (Modus, Typ, Original-ID). "
            "Die Rater-Dateien enthalten diese Metadaten nicht. "
            f"Fehlerhafte Generierungen sind {'im Scoring enthalten' if include_errors else 'vom Scoring ausgeschlossen'}. "
            "Zum Aggregieren der Rater-Bewertungen: experiments/aggregate_rater_scores.py --master <diese Datei>"
        ),
    )
    ws.cell(row=note_row, column=1).font = bold_font()
    ws.cell(row=note_row, column=2).alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{len(rows) + 1}"


def create_instruction_sheet(ws):
    ws.title = "Anleitung"

    ws.column_dimensions["A"].width = 120

    instructions = [
        "Bewertungshinweise",
        "Bitte bewerten Sie jede Antwort ausschließlich anhand der angezeigten Nutzereingabe",
        "Die Antworten wurden anonymisiert. Es ist nicht ersichtlich, ob eine Antwort mit oder ohne RAG erzeugt wurde",
        "Bewerten Sie die Qualität der Antwort, nicht deren Länge. Eine längere Antwort ist nicht automatisch besser",
        "Vergeben Sie für jedes Kriterium genau einen Wert von 1 bis 4",
        "",
        "Aufgabenpassung",
        "1 = Anforderungen kaum erfüllt oder verfehlt",
        "2 = teilweise passend, deutliche Abweichungen",
        "3 = überwiegend passend",
        "4 = vollständig und präzise passend",
        "",
        "Materialintegration",
        "1 = Materialien kaum, falsch oder irrelevant genutzt",
        "2 = Materialien vorhanden, aber oberflächlich genutzt",
        "3 = Materialien sinnvoll integriert",
        "4 = Materialien zentraler Bestandteil der Lösung",
        "Bei Eingaben ohne konkrete Materialien: Bewerten Sie, ob die Antwort plausibel mit der Materialunschärfe umgeht und keine unpassenden Annahmen trifft",
        "",
        "Umsetzbarkeit",
        "1 = nicht umsetzbar oder unverständlich",
        "2 = teilweise verständlich, deutliche Lücken",
        "3 = größtenteils umsetzbar",
        "4 = klar beschrieben und direkt umsetzbar",
        "",
        "Kreativität",
        "1 = sehr naheliegend oder trivial",
        "2 = leicht variierend, begrenzte Originalität",
        "3 = erkennbar kreativ",
        "4 = sehr originell und innovativ",
        "",
        "Nutzen Sie ausschließlich ganze Zahlen von 1 bis 4"
    ]

    for row, text in enumerate(instructions, start=1):
        cell = ws.cell(row=row, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True)

        if text in [
            "Bewertungshinweise",
            "Aufgabenpassung",
            "Materialintegration",
            "Umsetzbarkeit",
            "Kreativität"
        ]:
            cell.font = Font(bold=True)


def create_rater_workbook(rows, output_path, seed):
    wb = openpyxl.Workbook()

    ws_instructions = wb.active
    create_instruction_sheet(ws_instructions)

    ws = wb.create_sheet("Scoring")

    widths = {
        "A": 12,
        "B": 48,
        "C": 80,
        "D": 18,
        "E": 22,
        "F": 18,
        "G": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    headers = ["Antwort-ID", "Eingabe", "Antwort"] + DIMENSIONS

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font()
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

    scale_cell = ws.cell(
        row=2,
        column=4,
        value="Bewertung 1–4. Die vollständigen Kriterien befinden sich im Tabellenblatt 'Anleitung'"
    )
    scale_cell.font = Font(size=8, italic=True, color="666666")
    ws.merge_cells("D2:G2")

    scorable_rows = [row for row in rows if row["include_in_scoring"]]
    random.Random(seed).shuffle(scorable_rows)

    for row_idx, item in enumerate(scorable_rows, start=3):
        values = [
            item["response_id"],
            item["input"],
            item["response"],
            "", "", "", ""
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font()
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if col_idx in [4, 5, 6, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_idx].height = 95

    if scorable_rows:
        last_row = len(scorable_rows) + 2
        dv = DataValidation(
            type="whole",
            operator="between",
            formula1="1",
            formula2="4",
            allow_blank=True
        )
        dv.error = "Bitte nur ganze Zahlen von 1 bis 4 eintragen"
        dv.errorTitle = "Ungültige Bewertung"
        dv.prompt = "Bewertung von 1 bis 4 eingeben"
        dv.promptTitle = "Bewertungsskala"
        ws.add_data_validation(dv)
        dv.add(f"D3:G{last_row}")

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A1:G{len(scorable_rows) + 2}"

    wb.save(output_path)


def create_rater_files(rows, output_path, count, base_seed):
    if count <= 0:
        return []

    created_files = []
    for idx in range(1, count + 1):
        rater_path = output_path.with_name(
            f"{output_path.stem}_rater_{idx}{output_path.suffix}"
        )
        create_rater_workbook(rows, rater_path, seed=base_seed + idx)
        created_files.append(rater_path)

    return created_files


def main():
    parser = argparse.ArgumentParser(description="Experiment JSON -> anonymisierte Excel-Scoring-Datei")
    parser.add_argument("input", help="Path to experiment JSON file")
    parser.add_argument("--output", help="Output path for Excel file (optional)")
    parser.add_argument("--seed", type=int, default=42, help="Random seed for response order")
    parser.add_argument(
        "--include-errors",
        action="store_true",
        help="Include failed generations in the scoring sheet instead of excluding them",
    )
    parser.add_argument(
        "--rater-copies",
        type=int,
        default=4,
        help="Anzahl automatisch erzeugter Rater-Kopien. 0 deaktiviert die Kopien",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: file '{input_path}' not found")
        sys.exit(1)

    with open(input_path, encoding="utf-8") as f:
        data = json.load(f)

    results = data.get("results", [])
    if not results:
        print("Error: no results found in JSON file")
        sys.exit(1)

    rows = collect_responses(results, args.seed, include_errors=args.include_errors)
    scorable_count = sum(1 for row in rows if row["include_in_scoring"])

    if scorable_count == 0:
        print("Error: no successful/scorable responses found in JSON file")
        sys.exit(1)

    output_path = Path(args.output) if args.output else input_path.with_suffix(".xlsx")

    wb = openpyxl.Workbook()
    ws_mapping = wb.active

    create_mapping_sheet(ws_mapping, rows, args.seed, args.include_errors)

    wb.save(output_path)
    rater_files = create_rater_files(rows, output_path, args.rater_copies, args.seed)

    print(f"\nCreated: {output_path}")
    if rater_files:
        print("\nRater files:")
        for rater_file in rater_files:
            print(f"  - {rater_file}")


if __name__ == "__main__":
    main()
